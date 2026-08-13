"""
utils.py
UI Styling, Plotly Visualizations, and OpenPyXL Excel Export Utility Module.

Provides glassmorphic custom CSS styling with system font fallbacks for 100% offline usage,
Plotly dark-mode interactive charts, and openpyxl formatted Excel export generator.
"""

import io
from typing import List, Dict, Tuple, Optional
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def apply_custom_css():
    """
    Injects sleek modern custom CSS into Streamlit with offline system font fallbacks.
    """
    css = """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

    html, body, [class*="css"] {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
    }

    .main .block-container {
        padding-top: 1.5rem;
        padding-bottom: 3rem;
        max-width: 95%;
    }

    /* Glassmorphism Header Card */
    .galactic-header {
        background: linear-gradient(135deg, #0f172a 0%, #1e1b4b 50%, #312e81 100%);
        border-radius: 16px;
        padding: 2.2rem 2.5rem;
        color: #ffffff;
        box-shadow: 0 10px 30px rgba(0, 0, 0, 0.35);
        margin-bottom: 2rem;
        border: 1px solid rgba(255, 255, 255, 0.1);
    }
    .galactic-header h1 {
        font-size: 2.3rem;
        font-weight: 700;
        margin: 0 0 0.5rem 0;
        background: linear-gradient(90deg, #60a5fa, #a78bfa, #f472b6);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    .galactic-header p {
        font-size: 1.05rem;
        color: #cbd5e1;
        margin: 0;
    }

    /* Summary Metric Cards */
    .metric-card {
        border-radius: 12px;
        padding: 1.4rem 1.6rem;
        color: #ffffff;
        box-shadow: 0 4px 15px rgba(0, 0, 0, 0.2);
        transition: transform 0.2s ease, box-shadow 0.2s ease;
        margin-bottom: 1rem;
    }
    .metric-card:hover {
        transform: translateY(-3px);
        box-shadow: 0 8px 25px rgba(0, 0, 0, 0.25);
    }
    .card-good {
        background: linear-gradient(135deg, #059669 0%, #10b981 100%);
        border-left: 6px solid #34d399;
    }
    .card-moderate {
        background: linear-gradient(135deg, #d97706 0%, #f59e0b 100%);
        border-left: 6px solid #fbbf24;
    }
    .card-bad {
        background: linear-gradient(135deg, #dc2626 0%, #ef4444 100%);
        border-left: 6px solid #f87171;
    }
    .card-total {
        background: linear-gradient(135deg, #2563eb 0%, #3b82f6 100%);
        border-left: 6px solid #93c5fd;
    }
    .metric-card .card-title {
        font-size: 0.9rem;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        opacity: 0.9;
        margin-bottom: 0.3rem;
        font-weight: 600;
    }
    .metric-card .card-value {
        font-size: 2.2rem;
        font-weight: 700;
        margin: 0;
        line-height: 1.1;
    }
    .metric-card .card-subtitle {
        font-size: 0.85rem;
        opacity: 0.85;
        margin-top: 0.4rem;
    }
    </style>
    """
    st.markdown(css, unsafe_allow_html=True)


def plot_classification_pie(df: pd.DataFrame) -> go.Figure:
    """
    Creates a high-contrast, elegant Plotly Donut Chart showing classification breakdown.
    """
    if df.empty or "Result" not in df.columns:
        fig = go.Figure()
        fig.add_annotation(text="No Analysis Data Available", showarrow=False, font=dict(size=16, color="white"))
        fig.update_layout(template="plotly_dark", height=380)
        return fig

    counts = df["Result"].value_counts().to_dict()
    labels = ["GOOD", "MODERATE", "BAD"]
    values = [counts.get(l, 0) for l in labels]

    colors = {
        "GOOD": "#10b981",
        "MODERATE": "#f59e0b",
        "BAD": "#ef4444"
    }
    color_sequence = [colors[l] for l in labels]

    fig = go.Figure(data=[go.Pie(
        labels=labels,
        values=values,
        hole=0.55,
        marker=dict(colors=color_sequence, line=dict(color="#0f172a", width=2)),
        textinfo="label+percent",
        insidetextorientation="radial",
        hoverinfo="label+value+percent",
        textfont=dict(size=14, color="#ffffff", family="Inter, sans-serif")
    )])

    fig.update_layout(
        title=dict(text="<b>Classification Distribution</b>", font=dict(size=18, color="#f8fafc")),
        template="plotly_dark",
        height=380,
        margin=dict(l=20, r=20, t=50, b=20),
        legend=dict(orientation="h", yanchor="bottom", y=-0.1, xanchor="center", x=0.5),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)"
    )
    return fig


def plot_score_histogram(df: pd.DataFrame) -> go.Figure:
    """
    Creates an interactive Plotly Histogram showing Match Score distribution.
    """
    if df.empty or "Match Score" not in df.columns:
        fig = go.Figure()
        fig.add_annotation(text="No Score Data Available", showarrow=False, font=dict(size=16, color="white"))
        fig.update_layout(template="plotly_dark", height=380)
        return fig

    fig = px.histogram(
        df,
        x="Match Score",
        nbins=20,
        range_x=[0, 100],
        color="Result",
        color_discrete_map={
            "GOOD": "#10b981",
            "MODERATE": "#f59e0b",
            "BAD": "#ef4444"
        },
        labels={"Match Score": "Relevance Score (0 - 100)", "count": "Company Count"},
        title="<b>Relevance Score Distribution</b>"
    )

    fig.update_layout(
        template="plotly_dark",
        height=380,
        margin=dict(l=20, r=20, t=50, b=20),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(gridcolor="rgba(255,255,255,0.1)"),
        yaxis=dict(gridcolor="rgba(255,255,255,0.1)"),
        legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5)
    )
    return fig


def plot_top_keywords_bar(df: pd.DataFrame) -> go.Figure:
    """
    Extracts matched capabilities from the Reason column and plots a top 10 matched keywords bar chart.
    """
    if df.empty or "Reason" not in df.columns:
        fig = go.Figure()
        fig.add_annotation(text="No Keyword Data Available", showarrow=False, font=dict(size=16, color="white"))
        fig.update_layout(template="plotly_dark", height=380)
        return fig

    extracted_kws = []
    for reason in df["Reason"].dropna():
        if "Matched keywords:" in reason:
            part = reason.split("Matched keywords:")[1].split("|")[0].strip()
            if "(+" in part:
                part = part.split("(+")[0].strip()
            kws = [k.strip() for k in part.split(",") if k.strip()]
            extracted_kws.extend(kws)

    if not extracted_kws:
        fig = go.Figure()
        fig.add_annotation(text="No Specific Keyword Matches Recorded", showarrow=False, font=dict(size=16, color="white"))
        fig.update_layout(template="plotly_dark", height=380)
        return fig

    kw_series = pd.Series(extracted_kws).value_counts().head(10).sort_values(ascending=True)

    fig = px.bar(
        x=kw_series.values,
        y=kw_series.index,
        orientation='h',
        labels={'x': 'Company Match Frequency', 'y': 'Capabilities'},
        title="<b>Top Matched Manufacturing Capabilities</b>",
        color_discrete_sequence=["#60a5fa"]
    )

    fig.update_layout(
        template="plotly_dark",
        height=380,
        margin=dict(l=20, r=20, t=50, b=20),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(gridcolor="rgba(255,255,255,0.1)"),
        yaxis=dict(gridcolor="rgba(255,255,255,0.1)")
    )
    return fig


def generate_excel_download(df: pd.DataFrame) -> bytes:
    """
    Generates a beautifully formatted Excel file in memory using openpyxl.
    Applies color-coded fills for GOOD (green), MODERATE (amber), and BAD (red).
    """
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relevance Analysis"

    # Header styling
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # Classification Fills
    good_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    good_font = Font(name="Arial", size=10, bold=True, color="065F46")

    mod_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    mod_font = Font(name="Arial", size=10, bold=True, color="92400E")

    bad_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    bad_font = Font(name="Arial", size=10, bold=True, color="991B1B")

    # Write Headers
    columns = list(df.columns)
    ws.append(columns)

    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    # Write Rows
    result_col_idx = columns.index("Result") + 1 if "Result" in columns else -1

    for row_idx, row_data in enumerate(df.values, start=2):
        ws.append(list(row_data))
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = align_left

            if col_idx == result_col_idx:
                val = str(cell.value)
                if val == "GOOD":
                    cell.fill = good_fill
                    cell.font = good_font
                elif val == "MODERATE":
                    cell.fill = mod_fill
                    cell.font = mod_font
                elif val == "BAD":
                    cell.fill = bad_fill
                    cell.font = bad_font
                cell.alignment = align_center

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

    wb.save(output)
    return output.getvalue()


def generate_sample_dataset() -> pd.DataFrame:
    """
    Generates a realistic sample dataset of 20 companies with Cities for instant testing.
    """
    data = [
        {"co_name": "Apex Precision CNC Components", "category": "CNC Machining & Precision Parts", "keywords": "CNC milling, turning, aerospace precision components, tool room", "city": "Bengaluru", "website": "apexprecision.com", "email": "info@apexprecision.com"},
        {"co_name": "Titan Sheet Metal Tech", "category": "Sheet Metal Fabrication", "keywords": "Laser cutting, sheet metal bending, stamping, welding", "city": "Pune", "website": "titansheetmetal.io", "email": "contact@titansheetmetal.io"},
        {"co_name": "BioMed Plastics Ltd", "category": "Medical Injection Moulding", "keywords": "Medical devices, cleanroom injection moulding, precision plastics", "city": "Bengaluru", "website": "biomedplastics.org", "email": "sales@biomedplastics.org"},
        {"co_name": "AeroSpace Tech Dynamic", "category": "Aerospace & Defense Engineering", "keywords": "3D printing, titanium casting, forging, aerospace assembly", "city": "Bengaluru", "website": "aerospacetech.net", "email": "info@aerospacetech.net"},
        {"co_name": "Speedy Auto Dies & Molds", "category": "Automotive Tooling & Tool Room", "keywords": "Tool room, die casting, jigs fixture, automotive tooling", "city": "Chennai", "website": "speedyautodies.com", "email": "support@speedyautodies.com"},
        {"co_name": "NextGen Additive 3D", "category": "Additive Manufacturing & Prototyping", "keywords": "3D printing, rapid prototyping, SLA, SLS, metal 3D printing", "city": "Bengaluru", "website": "nextgen3d.com", "email": "hello@nextgen3d.com"},
        {"co_name": "Global Heavy Castings", "category": "Casting & Heavy Engineering", "keywords": "Sand casting, forging, heat treatment, heavy machinery parts", "city": "Mumbai", "website": "globalheavycastings.com", "email": "sales@globalheavycastings.com"},
        {"co_name": "LaserCut Precision Systems", "category": "Laser Cutting & Metal Stamping", "keywords": "Laser cutting, CNC punching, metal enclosure, wire EDM", "city": "Pune", "website": "lasercutprecision.de", "email": "info@lasercutprecision.de"},
        {"co_name": "Fresh Organic Farm Produce", "category": "Agriculture & Organic Food", "keywords": "Fresh vegetables, organic apples, dairy milk, farming", "city": "Delhi", "website": "freshorganicfarms.org", "email": "orders@freshorganicfarms.org"},
        {"co_name": "BlueSky Digital Marketing", "category": "Digital Marketing Agency", "keywords": "SEO, social media management, Google Ads, brand strategy", "city": "Mumbai", "website": "blueskymarketing.co", "email": "hello@blueskymarketing.co"},
        {"co_name": "Grand Horizon Boutique Hotel", "category": "Hospitality & Tourism", "keywords": "Luxury rooms, resort booking, spa, restaurant dining", "city": "Bengaluru", "website": "grandhorizonhotel.com", "email": "reservation@grandhorizonhotel.com"},
        {"co_name": "Metrology & Quality CMM Labs", "category": "Quality Inspection & Metrology", "keywords": "CMM inspection, quality assurance, metrology, precision measurement", "city": "Bengaluru", "website": "metrologycmmlabs.com", "email": "qa@metrologycmmlabs.com"},
        {"co_name": "PureWater Filtration Systems", "category": "Water Treatment & Filters", "keywords": "Reverse osmosis, industrial water filters, UV purification", "city": "Hyderabad", "website": "purewaterfilter.in", "email": "info@purewaterfilter.in"},
        {"co_name": "Swift Logistics & Freight", "category": "Logistics & Supply Chain", "keywords": "Warehouse storage, cargo transport, supply chain shipping", "city": "Mumbai", "website": "swiftlogistics.com", "email": "dispatch@swiftlogistics.com"},
        {"co_name": "Veritas Financial Planning", "category": "Financial Services & Accounting", "keywords": "Tax audit, wealth management, corporate accounting", "city": "Delhi", "website": "veritasfinance.com", "email": "contact@veritasfinance.com"},
        {"co_name": "Omega Injection Moulding Co", "category": "Plastic Moulding & Tooling", "keywords": "Injection moulding, plastic enclosures, rapid tooling", "city": "Chennai", "website": "omegamoulding.com", "email": "info@omegamoulding.com"},
        {"co_name": "CloudNova Software Solutions", "category": "SaaS & Cloud Computing", "keywords": "Python microservices, AWS architecture, web development", "city": "Bengaluru", "website": "cloudnova.io", "email": "dev@cloudnova.io"},
        {"co_name": "Vanguard Defence Gear", "category": "Defense & Military Equipment", "keywords": "Defense components, armaments machining, military hardware", "city": "Hyderabad", "website": "vanguarddefence.gov", "email": "contracts@vanguarddefence.gov"}
    ]
    return pd.DataFrame(data)
